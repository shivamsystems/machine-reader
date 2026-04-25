import easyocr
import re
import os
import time
from openpyxl import Workbook, load_workbook
from datetime import datetime
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

# ============ SETTINGS ============
EXCEL_FILE = "machine_readings.xlsx"
WATCH_FOLDER = "photos"
# ==================================

print("🔄 Loading OCR model...")
reader = easyocr.Reader(['en'], gpu=False, verbose=False)
print("✅ OCR model loaded!\n")


def fix_ocr_errors(text):
    """Fix common OCR character misreads."""
    
    # Convert to uppercase for consistency
    text = text.upper()
    
    # Fix common OCR mistakes in the KEY part (before the number value)
    replacements = [
        # Fix letter O to number 0
        ('O1K', '01K'), ('O2K', '02K'), ('O3K', '03K'), ('O4K', '04K'),
        ('O5K', '05K'), ('O6K', '06K'), ('O7K', '07K'), ('O8K', '08K'),
        ('O9K', '09K'),
        
        # Fix letter I to number 1
        ('I0K', '10K'), ('I1K', '11K'), ('I2K', '12K'), ('I3K', '13K'),
        ('I4K', '14K'), ('I5K', '15K'), ('I6K', '16K'),
        ('IOK', '10K'),  # IO = 10
        
        # Fix letter S to number 5
        ('OSK', '05K'), ('0SK', '05K'),
        ('ISK', '15K'), ('1SK', '15K'),
        
        # Fix M to 14 (common misread)
        (' MK', ' 14K'),
        
        # Fix letter L to number 1
        ('L0K', '10K'), ('L1K', '11K'), ('L2K', '12K'), ('L3K', '13K'),
        ('L4K', '14K'), ('L5K', '15K'), ('L6K', '16K'),
        
        # Fix other common issues
        ('IIK', '11K'),
        ('I K', '1K'),
        ('IKK', '11K'),
    ]
    
    for old, new in replacements:
        text = text.replace(old, new)
    
    return text


def extract_readings(text):
    """Extract all K readings from the text."""
    
    # First fix OCR errors
    fixed_text = fix_ocr_errors(text)
    
    print(f"\n   📝 After OCR fixes: {fixed_text[:150]}...")
    
    readings = {}
    
    # Method 1: Find pattern like "01K 76.15" or "01K: 76.15"
    # Matches: 1K, 01K, 1k, 01k followed by space/colon and number
    pattern = r'(\d{1,2})K[:\s]+(\d+\.?\d*)'
    
    matches = re.findall(pattern, fixed_text)
    
    for match in matches:
        key_num = int(match[0])
        if 1 <= key_num <= 16:  # Valid range 01K to 16K
            key = f"{str(key_num).zfill(2)}K"
            value = match[1]
            
            # Only update if we don't have this key, or this is a better value
            if key not in readings:
                readings[key] = value
    
    return readings, fixed_text


def read_and_save(image_path):
    """Extract all readings from image and save to Excel."""
    try:
        print(f"🔍 Processing image...")

        # Extract text from image
        results = reader.readtext(image_path)

        # Combine all detected text
        all_text = ""
        for (_, text, confidence) in results:
            all_text += " " + text

        print(f"   🔤 Raw OCR text:")
        print(f"   {all_text}")

        # Extract all K readings
        readings, fixed_text = extract_readings(all_text)
        
        # Sort and display
        print(f"\n   ✅ Found {len(readings)} readings:")
        for i in range(1, 17):
            key = f"{str(i).zfill(2)}K"
            value = readings.get(key, "❌ MISSING")
            print(f"   • {key}: {value}")

        # Check for missing readings
        missing = []
        for i in range(1, 17):
            key = f"{str(i).zfill(2)}K"
            if key not in readings:
                missing.append(key)
        
        if missing:
            print(f"\n   ⚠️  Missing readings: {', '.join(missing)}")
            print(f"   (Photo may need to be clearer)")

        # Save to Excel
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            # Create headers
            headers = ["Date", "Time"]
            for i in range(1, 17):
                headers.append(f"{str(i).zfill(2)}K")
            headers.extend(["Missing Count", "Raw OCR Text", "Image Name"])
            ws.append(headers)
            
            # Make header bold
            from openpyxl.styles import Font
            for cell in ws[1]:
                cell.font = Font(bold=True)

        # Prepare row data
        now = datetime.now()
        row = [
            now.strftime("%Y-%m-%d"),
            now.strftime("%H:%M:%S")
        ]
        
        # Add readings 01K to 16K
        for i in range(1, 17):
            key = f"{str(i).zfill(2)}K"
            value = readings.get(key, "")
            row.append(value)
        
        row.append(len(missing))  # Count of missing readings
        row.append(all_text[:500])  # Raw text for debugging
        row.append(os.path.basename(image_path))

        ws.append(row)
        wb.save(EXCEL_FILE)
        
        print(f"\n   💾 Saved to Excel!")
        print("=" * 60)

    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()


class PhotoHandler(FileSystemEventHandler):
    def on_created(self, event):
        if event.src_path.lower().endswith(('.jpg', '.jpeg', '.png', '.bmp')):
            print(f"\n📸 New photo: {os.path.basename(event.src_path)}")
            time.sleep(2)
            read_and_save(event.src_path)


def main():
    os.makedirs(WATCH_FOLDER, exist_ok=True)

    print("=" * 60)
    print("   MACHINE READING AUTO-PROCESSOR v2.0")
    print("   (Extracts 16 readings: 01K to 16K)")
    print("=" * 60)
    print(f"""
   📁 Drop photos in: {os.path.abspath(WATCH_FOLDER)}
   📊 Excel file: {EXCEL_FILE}
   
   Press Ctrl+C to stop
    """)
    print("👀 Waiting for photos...\n")

    observer = Observer()
    observer.schedule(PhotoHandler(), WATCH_FOLDER, recursive=False)
    observer.start()

    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
        print("\n🛑 Stopped!")

    observer.join()


if __name__ == "__main__":
    main()